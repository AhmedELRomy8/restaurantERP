"""
Reports & Analytics Screen
Sales summary, charts, and Excel/PDF export
"""
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime, timedelta
import database as db
from theme import COLORS, FONTS
from widgets import Toast


class ReportsScreen(tk.Frame):

    def __init__(self, parent, app, **kwargs):
        super().__init__(parent, bg=COLORS["bg_primary"], **kwargs)
        self.app = app
        self._build_ui()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        # ── Filter Bar ────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=COLORS["bg_secondary"], padx=16, pady=12)
        top.grid(row=0, column=0, sticky="ew")

        tk.Label(top, text="📊 التقارير والتحليلات",
                 fg=COLORS["accent"], bg=COLORS["bg_secondary"],
                 font=FONTS["title_md"]).pack(side="left")

        # Quick ranges
        self.range_var = tk.StringVar(value="today")
        for val, lbl in [("today", "اليوم"), ("week", "أسبوع"), ("month", "الشهر"), ("custom", "مخصص")]:
            tk.Radiobutton(top, text=lbl, variable=self.range_var, value=val,
                           bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"],
                           selectcolor=COLORS["accent"], activebackground=COLORS["bg_secondary"],
                           font=FONTS["body_sm"], indicatoron=False,
                           relief="flat", padx=10, pady=5, cursor="hand2",
                           command=self._on_range_change).pack(side="right", padx=2)

        tk.Label(top, text="الفترة:", fg=COLORS["text_secondary"],
                 bg=COLORS["bg_secondary"], font=FONTS["body_sm"]).pack(side="right", padx=8)

        # Export buttons
        tk.Button(top, text="📥 تصدير Excel",
                  bg=COLORS["success"], fg="white",
                  font=FONTS["body_sm"], bd=0, padx=12, pady=6, relief="flat",
                  cursor="hand2", command=self._export_excel).pack(side="left", padx=(16, 4))

        tk.Button(top, text="🖨️ تصدير PDF",
                  bg=COLORS["info"], fg="white",
                  font=FONTS["body_sm"], bd=0, padx=12, pady=6, relief="flat",
                  cursor="hand2", command=self._export_pdf).pack(side="left", padx=4)

        # ── Custom Date Row ────────────────────────────────────────────────────
        self.date_frame = tk.Frame(self, bg=COLORS["bg_secondary"], padx=16, pady=6)
        self.date_frame.grid(row=1, column=0, sticky="ew")

        from tkcalendar import DateEntry
        tk.Label(self.date_frame, text="من:", fg=COLORS["text_secondary"],
                 bg=COLORS["bg_secondary"], font=FONTS["body_sm"]).pack(side="left")
        self.date_from = DateEntry(self.date_frame, font=FONTS["body_sm"],
                                    background=COLORS["accent"], foreground="black",
                                    borderwidth=0, date_pattern="yyyy-mm-dd")
        self.date_from.pack(side="left", padx=6)

        tk.Label(self.date_frame, text="إلى:", fg=COLORS["text_secondary"],
                 bg=COLORS["bg_secondary"], font=FONTS["body_sm"]).pack(side="left", padx=(12, 0))
        self.date_to = DateEntry(self.date_frame, font=FONTS["body_sm"],
                                  background=COLORS["accent"], foreground="black",
                                  borderwidth=0, date_pattern="yyyy-mm-dd")
        self.date_to.pack(side="left", padx=6)

        tk.Button(self.date_frame, text="🔍 عرض", bg=COLORS["accent"], fg="black",
                  font=FONTS["body_sm"], bd=0, padx=14, pady=5, relief="flat",
                  cursor="hand2", command=self._load).pack(side="left", padx=10)

        # ── Content ───────────────────────────────────────────────────────────
        content = tk.Frame(self, bg=COLORS["bg_primary"])
        content.grid(row=2, column=0, sticky="nsew", padx=16, pady=8)
        content.columnconfigure(0, weight=1)
        content.columnconfigure(1, weight=1)
        content.rowconfigure(1, weight=1)
        self.rowconfigure(2, weight=1)

        # KPI row
        kpi_row = tk.Frame(content, bg=COLORS["bg_primary"])
        kpi_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        kpi_row.columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.kpi = {}
        kpis = [
            ("total_orders",   "إجمالي الطلبات",  "📋", COLORS["info"]),
            ("total_revenue",  "إجمالي الإيرادات","💰", COLORS["success"]),
            ("avg_order",      "متوسط الطلب",     "📈", COLORS["accent"]),
            ("total_discount", "إجمالي الخصومات", "🏷️", COLORS["warning"]),
            ("total_tax",      "إجمالي الضريبة",  "🧾", COLORS["danger"]),
        ]
        for i, (key, title, icon, color) in enumerate(kpis):
            card = self._kpi_card(kpi_row, title, "—", icon, color)
            card.grid(row=0, column=i, sticky="ew", padx=4)
            self.kpi[key] = card

        # Payment breakdown (left)
        pay_frame = tk.Frame(content, bg=COLORS["bg_card"], bd=0,
                              highlightthickness=1, highlightbackground=COLORS["border"],
                              padx=16, pady=12)
        pay_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 6))

        tk.Label(pay_frame, text="💳 توزيع طرق الدفع",
                 fg=COLORS["accent"], bg=COLORS["bg_card"],
                 font=FONTS["title_sm"]).pack(anchor="w", pady=(0, 12))

        self.pay_canvas = tk.Canvas(pay_frame, bg=COLORS["bg_card"],
                                     height=160, highlightthickness=0)
        self.pay_canvas.pack(fill="x")

        self.pay_labels = tk.Frame(pay_frame, bg=COLORS["bg_card"])
        self.pay_labels.pack(fill="x", pady=8)

        # Daily chart (right)
        chart_frame = tk.Frame(content, bg=COLORS["bg_card"], bd=0,
                                highlightthickness=1, highlightbackground=COLORS["border"],
                                padx=16, pady=12)
        chart_frame.grid(row=1, column=1, sticky="nsew", padx=(6, 0))

        tk.Label(chart_frame, text="📅 المبيعات اليومية",
                 fg=COLORS["accent"], bg=COLORS["bg_card"],
                 font=FONTS["title_sm"]).pack(anchor="w", pady=(0, 8))

        self.daily_canvas = tk.Canvas(chart_frame, bg=COLORS["bg_card"],
                                       highlightthickness=0)
        self.daily_canvas.pack(fill="both", expand=True)

        # Top items table
        items_frame = tk.Frame(content, bg=COLORS["bg_card"], bd=0,
                                highlightthickness=1, highlightbackground=COLORS["border"],
                                padx=16, pady=12)
        items_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(8, 0))
        content.rowconfigure(2, weight=1)

        tk.Label(items_frame, text="🏆 الأصناف الأكثر مبيعاً",
                 fg=COLORS["accent"], bg=COLORS["bg_card"],
                 font=FONTS["title_sm"]).pack(anchor="w", pady=(0, 8))

        # Headers
        hdr = tk.Frame(items_frame, bg=COLORS["bg_hover"])
        hdr.pack(fill="x")
        for col, w in [("#", 40), ("الصنف", 200), ("الفئة", 120),
                        ("الكمية", 80), ("الإيراد", 120)]:
            tk.Label(hdr, text=col, fg=COLORS["accent"],
                     bg=COLORS["bg_hover"], font=FONTS["body_sm"],
                     width=w//8, pady=6).pack(side="left", padx=4)

        scroll_frame = tk.Frame(items_frame, bg=COLORS["bg_card"])
        scroll_frame.pack(fill="both", expand=True)

        self.items_canvas = tk.Canvas(scroll_frame, bg=COLORS["bg_card"], highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_frame, command=self.items_canvas.yview)
        self.items_inner = tk.Frame(self.items_canvas, bg=COLORS["bg_card"])
        self.items_inner.bind("<Configure>",
                               lambda e: self.items_canvas.configure(
                                   scrollregion=self.items_canvas.bbox("all")))
        self.items_canvas.create_window((0, 0), window=self.items_inner, anchor="nw")
        self.items_canvas.configure(yscrollcommand=scrollbar.set)
        self.items_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self._on_range_change()

    def _kpi_card(self, parent, title, value, icon, color):
        f = tk.Frame(parent, bg=COLORS["bg_card"], bd=0,
                     highlightthickness=1, highlightbackground=COLORS["border"],
                     padx=12, pady=10)
        bar = tk.Frame(f, bg=color, width=4)
        bar.pack(side="left", fill="y")
        body = tk.Frame(f, bg=COLORS["bg_card"], padx=10)
        body.pack(side="left", fill="both", expand=True)
        tk.Label(body, text=f"{icon} {title}", fg=COLORS["text_secondary"],
                 bg=COLORS["bg_card"], font=FONTS["body_sm"]).pack(anchor="w")
        f.val = tk.Label(body, text=value, fg=color,
                          bg=COLORS["bg_card"], font=FONTS["title_sm"])
        f.val.pack(anchor="w", pady=(2, 0))
        return f

    def _on_range_change(self):
        today = datetime.now()
        r = self.range_var.get()
        if r == "today":
            d_from = today.strftime("%Y-%m-%d")
            d_to   = today.strftime("%Y-%m-%d")
        elif r == "week":
            d_from = (today - timedelta(days=6)).strftime("%Y-%m-%d")
            d_to   = today.strftime("%Y-%m-%d")
        elif r == "month":
            d_from = today.strftime("%Y-%m-01")
            d_to   = today.strftime("%Y-%m-%d")
        else:
            return  # custom: user uses date pickers

        self.date_from.set_date(datetime.strptime(d_from, "%Y-%m-%d"))
        self.date_to.set_date(datetime.strptime(d_to, "%Y-%m-%d"))
        self._load()

    def _load(self):
        d_from = self.date_from.get_date().strftime("%Y-%m-%d")
        d_to   = self.date_to.get_date().strftime("%Y-%m-%d")
        data   = db.get_report_data(d_from, d_to)
        s = data["summary"]

        orders = s.get("total_orders", 0)
        revenue = s.get("total_revenue", 0)
        avg = (revenue / orders) if orders > 0 else 0

        self.kpi["total_orders"].val.config(text=str(orders))
        self.kpi["total_revenue"].val.config(text=f"{revenue:.2f} ج.م")
        self.kpi["avg_order"].val.config(text=f"{avg:.2f} ج.م")
        self.kpi["total_discount"].val.config(text=f"{s.get('total_discount', 0):.2f} ج.م")
        self.kpi["total_tax"].val.config(text=f"{s.get('total_tax', 0):.2f} ج.م")

        self._draw_payment_chart(s)
        self._draw_daily_chart(data["daily"])
        self._draw_items(data["items"])

    def _draw_payment_chart(self, s):
        c = self.pay_canvas
        c.delete("all")
        c.update_idletasks()

        vals = {
            "نقدي 💵":   (s.get("cash_revenue", 0),   COLORS["success"]),
            "بطاقة 💳":  (s.get("card_revenue", 0),   COLORS["info"]),
            "محفظة 📱":  (s.get("wallet_revenue", 0), COLORS["accent"]),
        }
        total = sum(v for v, _ in vals.values()) or 1

        # Draw pie-like bar chart
        for w in self.pay_labels.winfo_children():
            w.destroy()

        y = 10
        for label, (val, color) in vals.items():
            pct = val / total
            bar_w = int(pct * (c.winfo_width() or 300))
            c.create_rectangle(0, y, bar_w, y + 34, fill=color, outline="")
            c.create_text(max(bar_w // 2, 60), y + 17,
                           text=f"{label}: {val:.0f} ج.م ({pct*100:.0f}%)",
                           fill="white", font=("Cairo", 10, "bold"))
            y += 44

        # Legend
        for label, (val, color) in vals.items():
            row = tk.Frame(self.pay_labels, bg=COLORS["bg_card"])
            row.pack(side="left", padx=12)
            tk.Frame(row, bg=color, width=12, height=12).pack(side="left")
            tk.Label(row, text=f" {label}: {val:.0f} ج.م",
                     fg=COLORS["text_secondary"], bg=COLORS["bg_card"],
                     font=FONTS["body_sm"]).pack(side="left")

    def _draw_daily_chart(self, daily):
        c = self.daily_canvas
        c.delete("all")
        if not daily:
            c.create_text(200, 80, text="لا توجد بيانات",
                           fill=COLORS["text_muted"], font=FONTS["body_sm"])
            return

        c.update_idletasks()
        w = c.winfo_width() or 400
        h = c.winfo_height() or 200
        pad_l, pad_b, pad_r, pad_t = 40, 30, 20, 20

        max_rev = max(r["revenue"] for r in daily) or 1
        bar_w   = max(10, (w - pad_l - pad_r) // len(daily) - 4)

        for i, row in enumerate(daily):
            x = pad_l + i * (bar_w + 4)
            bar_h = int((row["revenue"] / max_rev) * (h - pad_b - pad_t))
            y1, y2 = h - pad_b - bar_h, h - pad_b
            # Gradient effect (two rectangles)
            c.create_rectangle(x, y1, x + bar_w, y2,
                                fill=COLORS["accent"], outline="")
            c.create_rectangle(x, y1, x + bar_w, y1 + 4,
                                fill=COLORS["accent_light"], outline="")
            # Date label
            day = row["day"][-5:]  # MM-DD
            c.create_text(x + bar_w // 2, h - pad_b + 12,
                           text=day, fill=COLORS["text_muted"],
                           font=("Cairo", 7), angle=0)
            # Value label
            if bar_h > 20:
                c.create_text(x + bar_w // 2, y1 - 8,
                               text=f"{row['revenue']:.0f}",
                               fill=COLORS["text_secondary"], font=("Cairo", 7))

    def _draw_items(self, items):
        for w in self.items_inner.winfo_children():
            w.destroy()

        for i, item in enumerate(items[:20]):
            bg = COLORS["bg_hover"] if i % 2 else COLORS["bg_card"]
            row = tk.Frame(self.items_inner, bg=bg)
            row.pack(fill="x", pady=1)
            for text, wid in [
                (str(i + 1),                          40),
                (item["name"],                        200),
                (item.get("category", "—"),           120),
                (str(item.get("total_qty", 0)),        80),
                (f"{item.get('total_revenue', 0):.2f} ج.م", 120),
            ]:
                tk.Label(row, text=text, fg=COLORS["text_primary"],
                         bg=bg, font=FONTS["body_sm"],
                         width=wid // 8, pady=6).pack(side="left", padx=4)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            Toast(self, "مكتبة openpyxl غير متوفرة", "error")
            return

        d_from = self.date_from.get_date().strftime("%Y-%m-%d")
        d_to   = self.date_to.get_date().strftime("%Y-%m-%d")
        data   = db.get_report_data(d_from, d_to)

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"تقرير_{d_from}_{d_to}.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ملخص التقرير"
        ws.sheet_view.rightToLeft = True

        header_fill  = PatternFill("solid", fgColor="F0A500")
        header_font  = Font(name="Cairo", bold=True, color="000000", size=12)
        title_font   = Font(name="Cairo", bold=True, color="F0A500", size=14)
        center_align = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:E1")
        ws["A1"] = f"تقرير المبيعات  ({d_from} → {d_to})"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 30

        s = data["summary"]
        summary_data = [
            ("إجمالي الطلبات", s.get("total_orders", 0)),
            ("إجمالي الإيرادات (ج.م)", s.get("total_revenue", 0)),
            ("إجمالي الخصومات (ج.م)", s.get("total_discount", 0)),
            ("إجمالي الضريبة (ج.م)", s.get("total_tax", 0)),
            ("إيرادات نقدي (ج.م)", s.get("cash_revenue", 0)),
            ("إيرادات بطاقة (ج.م)", s.get("card_revenue", 0)),
            ("إيرادات محفظة (ج.م)", s.get("wallet_revenue", 0)),
        ]
        for r, (k, v) in enumerate(summary_data, start=3):
            ws.cell(r, 1, k).font = Font(name="Cairo", size=11)
            ws.cell(r, 2, v).font = Font(name="Cairo", size=11)
            ws.cell(r, 1).alignment = Alignment(horizontal="right")
            ws.cell(r, 2).alignment = Alignment(horizontal="center")

        # Items sheet
        ws2 = wb.create_sheet("الأصناف")
        ws2.sheet_view.rightToLeft = True
        headers = ["الصنف", "الفئة", "الكمية المباعة", "الإيراد (ج.م)"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(1, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            ws2.column_dimensions[chr(64 + c)].width = 20

        for r, item in enumerate(data["items"], 2):
            for c, val in enumerate([
                item["name"],
                item.get("category", ""),
                item.get("total_qty", 0),
                round(item.get("total_revenue", 0), 2),
            ], 1):
                ws2.cell(r, c, val).alignment = Alignment(horizontal="center")

        wb.save(path)
        Toast(self, f"تم التصدير: {path.split('/')[-1]}", "success")

    def _export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as pdf_canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib import colors as rc
        except ImportError:
            Toast(self, "مكتبة reportlab غير متوفرة", "error")
            return

        d_from = self.date_from.get_date().strftime("%Y-%m-%d")
        d_to   = self.date_to.get_date().strftime("%Y-%m-%d")
        data   = db.get_report_data(d_from, d_to)

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"report_{d_from}_{d_to}.pdf"
        )
        if not path:
            return

        try:
            import os
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Cairo-Regular.ttf")
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("Cairo", font_path))
                font_name = "Cairo"
            else:
                font_name = "Helvetica"
        except Exception:
            font_name = "Helvetica"

        c = pdf_canvas.Canvas(path, pagesize=A4)
        w_pt, h_pt = A4
        c.setFont(font_name, 16)
        c.setFillColor(rc.HexColor("#F0A500"))
        c.drawCentredString(w_pt / 2, h_pt - 60, f"Sales Report  {d_from} to {d_to}")

        c.setFont(font_name, 11)
        c.setFillColor(rc.black)
        s = data["summary"]
        y = h_pt - 100
        lines = [
            f"Total Orders: {s.get('total_orders', 0)}",
            f"Total Revenue: {s.get('total_revenue', 0):.2f} EGP",
            f"Total Discounts: {s.get('total_discount', 0):.2f} EGP",
            f"Total Tax: {s.get('total_tax', 0):.2f} EGP",
            f"Cash: {s.get('cash_revenue', 0):.2f} EGP",
            f"Card: {s.get('card_revenue', 0):.2f} EGP",
            f"Wallet: {s.get('wallet_revenue', 0):.2f} EGP",
        ]
        for line in lines:
            c.drawString(60, y, line)
            y -= 22

        y -= 20
        c.setFont(font_name, 13)
        c.setFillColor(rc.HexColor("#F0A500"))
        c.drawString(60, y, "Top Items:")
        y -= 20
        c.setFont(font_name, 10)
        c.setFillColor(rc.black)
        for i, item in enumerate(data["items"][:15], 1):
            c.drawString(60, y, f"{i}. {item['name']}  - Qty: {item.get('total_qty', 0)}  "
                                  f"- Revenue: {item.get('total_revenue', 0):.2f} EGP")
            y -= 18
            if y < 60:
                c.showPage()
                y = h_pt - 60

        c.save()
        Toast(self, f"تم التصدير: {path.split('/')[-1]}", "success")

    def on_show(self):
        self._on_range_change()
